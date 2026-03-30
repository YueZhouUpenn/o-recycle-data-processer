#!/usr/bin/env python3
"""
生成测试数据集
100条出库数据：30条未回收，30条现场回收，30条统一回收，10条退货
"""

from openpyxl import Workbook
from datetime import datetime, timedelta
import random

# 基础数据
products = [
    ("P020400286", "点式治疗头", "D4.5"),
    ("P020400422", "手柄", "MFUS M3.0"),
    ("P020400330", "治疗头", "D-4.5"),
    ("P020400011", "治疗头", "D3.0"),
]

salespeople = ["张三", "李四", "王五", "赵六", "钱七"]
depts = ["西北区", "西南区", "华东区", "华北区", "华南区"]
hospitals = [
    "西安第一人民医院", "成都锦欣健康管理有限公司", "山东凯奥医疗科技有限公司",
    "南京嘉会国际医院", "上海仁济医院", "北京协和医院", "广州中山大学附属第一医院"
]

# 生成序列号
def gen_serial(i):
    return f"US{str(i).zfill(2)}CS{str(i*2).zfill(2)}GG{str(i*3).zfill(3)}"

# 生成出库单
def generate_outbound():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # 表头
    headers = ["序列号", "物料编码", "物料名称", "规格型号", "日期", "单据编号", 
               "订单单号", "销售员", "销售部门", "客户", "终端客户", "单据类型", "描述"]
    ws.append(headers)
    
    # 生成100条数据
    base_date = datetime(2025, 1, 1)
    for i in range(1, 101):
        product = random.choice(products)
        salesperson = random.choice(salespeople)
        dept = random.choice(depts)
        hospital = random.choice(hospitals)
        
        date = (base_date + timedelta(days=i*3)).strftime("%Y/%m/%d")
        doc_no = f"XSCK{100 + i:012d}"
        order_no = f"XSDD{1000 + i:013d}"
        
        row = [
            gen_serial(i),
            product[0],
            product[1],
            product[2],
            date,
            doc_no,
            order_no,
            salesperson,
            dept,
            hospital,
            hospital,
            "标准销售出库单",
            f"{product[2]},含包装"
        ]
        ws.append(row)
    
    wb.save("input_test/销售出库单.xlsx")
    print("✓ 已生成 销售出库单.xlsx (100条)")

# 生成现场回收（30条，序列号1-30）
def generate_onsite_recycle():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    headers = ["序列号", "状态", "剩余发数", "实际回收客户", "现场实际回收日期", "现场实际回收运单号",
               "单据编码（必填）", "规格型号", "回收方式", "扫码回收（必填）", "折扣订单指令号", "扫码",
               "运单单号", "产品编码", "ERP指令号", "终端医院", "客户名称", "实际回收终端", "描述",
               "回收序列号", "产品名称", "销售订单", "销售出库单", "锁定状态", "回收日期", "出库日期",
               "ERP出库单号", "创建人", "创建时间", "业务类型", "归属部门", "负责人主属部门",
               "负责人（必填）", "生命状态", "销售经理", "来源", "最后修改人", "最后修改时间",
               "版本", "所属集团", "下单日期", "终端机构"]
    ws.append(headers)
    
    base_date = datetime(2025, 6, 1)
    for i in range(1, 31):
        product = random.choice(products)
        hospital = random.choice(hospitals)
        salesperson = random.choice(salespeople)
        dept = random.choice(depts)
        
        recycle_date = (base_date + timedelta(days=i*5)).strftime("%Y-%m-%d")
        
        row = [
            gen_serial(i),  # 序列号
            "待处理",
            0,
            hospital,
            recycle_date,
            f"SF{1000000 + i:011d}",
            f"SMMX{20250600 + i:08d}",
            product[2],
            "现场报废",
            f"SMHS{20250600 + i:08d}",
            "",
            f"{product[0]}*{gen_serial(i)}",
            "",
            product[0],
            f"XSDD{1000 + i:013d}",
            hospital,
            hospital,
            hospital,
            product[2],
            gen_serial(i),
            product[1],
            "",
            "",
            "未锁定",
            recycle_date,
            (base_date - timedelta(days=180 + i)).strftime("%Y-%m-%d"),
            f"XSCK{100 + i:012d}",
            salesperson,
            recycle_date + " 15:30",
            "预设业务类型",
            dept,
            dept,
            salesperson,
            "正常",
            salesperson,
            "销售出库",
            "系统",
            recycle_date + " 16:00",
            "黑钻",
            "集团公司",
            "",
            hospital
        ]
        ws.append(row)
    
    wb.save("input_test/现场回收.xlsx")
    print("✓ 已生成 现场回收.xlsx (30条，序列号1-30)")

# 生成统一回收（30条，序列号31-60）
def generate_unified_recycle():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    headers = ["序列号", "状态", "剩余发数", "实际回收客户", "现场实际回收日期", "现场实际回收运单号",
               "单据编码（必填）", "规格型号", "回收方式", "扫码回收（必填）", "折扣订单指令号", "扫码",
               "运单单号", "产品编码", "ERP指令号", "终端医院", "客户名称", "实际回收终端", "描述",
               "回收序列号", "产品名称", "销售订单", "销售出库单", "锁定状态", "回收日期", "出库日期",
               "ERP出库单号", "创建人", "创建时间", "业务类型", "归属部门", "负责人主属部门",
               "负责人（必填）", "生命状态", "销售经理", "来源", "最后修改人", "最后修改时间",
               "版本", "所属集团", "下单日期", "终端机构"]
    ws.append(headers)
    
    base_date = datetime(2025, 7, 1)
    for i in range(31, 61):
        product = random.choice(products)
        hospital = random.choice(hospitals)
        salesperson = random.choice(salespeople)
        dept = random.choice(depts)
        
        recycle_date = (base_date + timedelta(days=i*4)).strftime("%Y-%m-%d")
        
        row = [
            gen_serial(i),
            "待处理",
            0,
            hospital,
            "",  # 现场实际回收日期（统一回收为空）
            "",  # 现场实际回收运单号（统一回收为空）
            f"SMMX{20250700 + i:08d}",
            product[2],
            "统一报废",
            f"SMHS{20250700 + i:08d}",
            "",
            f"{product[0]}*{gen_serial(i)}",
            f"SF{2000000 + i:011d}",  # 运单单号（统一回收用这个）
            product[0],
            f"XSDD{1000 + i:013d}",
            hospital,
            hospital,
            hospital,
            product[2],
            gen_serial(i),
            product[1],
            "",
            "",
            "未锁定",
            recycle_date,
            (base_date - timedelta(days=200 + i)).strftime("%Y-%m-%d"),
            f"XSCK{100 + i:012d}",
            salesperson,
            recycle_date + " 10:20",
            "预设业务类型",
            dept,
            dept,
            salesperson,
            "审核中",
            salesperson,
            "销售出库",
            "系统",
            recycle_date + " 11:00",
            "黄金版3.0GX",
            "集团公司",
            "",
            hospital
        ]
        ws.append(row)
    
    wb.save("input_test/统一回收.xlsx")
    print("✓ 已生成 统一回收.xlsx (30条，序列号31-60)")

# 生成退货表（10条，序列号61-70）
def generate_return():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    headers = ["类别", "库存方向", "序列号", "物料编码", "物料名称", "规格型号", "日期",
               "单据编号", "指令号", "领料人", "领料部门", "客户", "终端客户", "其他出库类型", "收货地址"]
    ws.append(headers)
    
    base_date = datetime(2025, 8, 1)
    for i in range(61, 71):
        product = random.choice(products)
        hospital = random.choice(hospitals)
        salesperson = random.choice(salespeople)
        
        return_date = (base_date + timedelta(days=i*2)).strftime("%Y/%m/%d")
        
        row = [
            "一代",
            "退货",
            gen_serial(i),
            product[0],
            product[1],
            product[2],
            return_date,
            f"QTCK{1000 + i:012d}",
            "",
            salesperson,
            "售后服务部/东区",
            hospital,
            "",
            "C类（保内更换）",
            ""
        ]
        ws.append(row)
    
    wb.save("input_test/退货表.xlsx")
    print("✓ 已生成 退货表.xlsx (10条，序列号61-70)")

if __name__ == "__main__":
    print("开始生成测试数据...")
    generate_outbound()
    generate_onsite_recycle()
    generate_unified_recycle()
    generate_return()
    print("\n✓ 测试数据生成完成！")
    print("  - 100条出库数据")
    print("  - 30条现场回收（序列号1-30）")
    print("  - 30条统一回收（序列号31-60）")
    print("  - 10条退货（序列号61-70）")
    print("  - 30条未回收（序列号71-100）")
